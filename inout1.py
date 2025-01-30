import tkinter as tk
from tkinter import filedialog, ttk
import openpyxl

def transfer_data():
    out_path = out_entry.get()
    in_path = in_entry.get()
    out_sheet_name = out_sheet_entry.get()  # 転送元シート名
    in_sheet_name = in_sheet_entry.get()  # 転送先シート名

    try:
        out_wb = openpyxl.load_workbook(out_path)
        in_wb = openpyxl.load_workbook(in_path)

        if out_sheet_name not in out_wb.sheetnames or in_sheet_name not in in_wb.sheetnames:
            raise ValueError("指定されたシートが見つかりません")

        out_sheet = out_wb[out_sheet_name]  # シートを指定
        in_sheet = in_wb[in_sheet_name]

        # 転記処理 (例)
        transfer_list = [
            ("G5", "J7"),  # out.xlsxのG5セル -> in.xlsxのJ7セル
            ("H6", "K8"),  # out.xlsxのH6セル -> in.xlsxのK8セル
            ("I7", "L9"),  # out.xlsxのI7セル -> in.xlsxのL9セル
            ("J8", "M10"), # out.xlsxのJ8セル -> in.xlsxのM10セル
            ("K9", "N11"), # out.xlsxのK9セル -> in.xlsxのN11セル
        ]

        for out_cell, in_cell in transfer_list:
            value = out_sheet[out_cell].value
            in_sheet[in_cell].value = value

        in_wb.save(in_path)  # in.xlsxを保存
        result_label.config(text="転記完了")

    except FileNotFoundError:
        result_label.config(text="ファイルが見つかりません")
    except ValueError as e:
        result_label.config(text=f"エラー: {e}")
    except Exception as e:
        result_label.config(text=f"エラー: {e}")

root = tk.Tk()
root.title("Excelデータ転送")

# ファイルパス入力
out_label = tk.Label(root, text="out.xlsxパス:")
out_label.grid(row=0, column=0)
out_entry = tk.Entry(root)
out_entry.grid(row=0, column=1)
out_button = tk.Button(root, text="参照", command=lambda: out_entry.insert(0, filedialog.askopenfilename()))
out_button.grid(row=0, column=2)

in_label = tk.Label(root, text="in.xlsxパス:")
in_label.grid(row=1, column=0)
in_entry = tk.Entry(root)
in_entry.grid(row=1, column=1)
in_button = tk.Button(root, text="参照", command=lambda: in_entry.insert(0, filedialog.askopenfilename()))
in_button.grid(row=1, column=2)

# シート名入力
out_sheet_label = tk.Label(root, text="out.xlsxシート名:")
out_sheet_label.grid(row=2, column=0)
out_sheet_entry = tk.Entry(root)
out_sheet_entry.grid(row=2, column=1)

in_sheet_label = tk.Label(root, text="in.xlsxシート名:")
in_sheet_label.grid(row=3, column=0)
in_sheet_entry = tk.Entry(root)
in_sheet_entry.grid(row=3, column=1)

# 転送ボタン
transfer_button = tk.Button(root, text="転送", command=transfer_data)
transfer_button.grid(row=4, column=0, columnspan=3)

# 結果表示ラベル
result_label = tk.Label(root, text="")
result_label.grid(row=5, column=0, columnspan=3)

root.mainloop()